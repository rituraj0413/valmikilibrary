import base64
import hashlib
import hmac
import json
from datetime import date
from decimal import Decimal, InvalidOperation
from functools import wraps
from urllib.parse import urlencode
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core import mail, signing
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Max, Sum
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_time
from django.views.decorators.http import require_POST
from openpyxl.styles import Font

from .finance import (
    MONTH_NAMES,
    build_ledger_row,
    get_month_status,
    month_label,
    monthly_reminder_rows,
    monthly_report_rows,
    overdue_summary,
    paid_lookup,
    yearly_report_rows,
)
from .models import Mentor, MentorshipSession, Payment, PortalMessage, Student, StudentOnlinePayment
from .pdf_utils import build_simple_pdf

STUDENT_PORTAL_SESSION_KEY = 'student_portal_student_id'
STUDENT_PORTAL_SIGNING_SALT = 'student-portal-verification'
MENTOR_PORTAL_SESSION_KEY = 'mentor_portal_mentor_id'
MENTOR_PORTAL_SIGNING_SALT = 'mentor-portal-verification'
SHIFT_DISPLAY_ORDER = {'Full Day': 0, 'Morning': 1, 'Evening': 2}
PUBLIC_USE_CASES = [
    {
        'icon': 'bi bi-book-half',
        'title': 'Private Libraries',
        'description': 'Manage seat allocation, fees, receipts, and daily operations from one branded dashboard.',
        'accent': 'emerald',
    },
    {
        'icon': 'bi bi-laptop',
        'title': 'Study Rooms',
        'description': 'Track shift-based seating, occupancy, reminders, and collections with zero manual confusion.',
        'accent': 'sky',
    },
    {
        'icon': 'bi bi-mortarboard-fill',
        'title': 'Coaching Institutes',
        'description': 'Combine library access, mentorship, student profiles, and fee control in one workflow.',
        'accent': 'gold',
    },
    {
        'icon': 'bi bi-buildings-fill',
        'title': 'School & College Libraries',
        'description': 'Bring structure to reading spaces, memberships, reporting, and student engagement.',
        'accent': 'violet',
    },
]
PUBLIC_FEATURES = [
    {
        'icon': 'bi bi-grid-1x2-fill',
        'title': 'Live Seat & Shift Management',
        'description': 'See which seats are vacant, partially occupied, or full day occupied with clear visual control.',
    },
    {
        'icon': 'bi bi-cash-stack',
        'title': 'Payment Ledger & Receipts',
        'description': 'Mark fees paid, auto-calculate dues, export receipts, and maintain a clean payment history.',
    },
    {
        'icon': 'bi bi-person-vcard-fill',
        'title': 'Student Profiles & Portal',
        'description': 'Give every student a profile with personal, academic, and seat information plus receipt access.',
    },
    {
        'icon': 'bi bi-megaphone-fill',
        'title': 'Smart Reminders & Dues Follow-Up',
        'description': 'Track pending months, prepare monthly reminder messages, and reduce collection delays.',
    },
    {
        'icon': 'bi bi-file-earmark-bar-graph-fill',
        'title': 'Reports for Owners',
        'description': 'Export monthly and yearly reports in Excel or PDF to review operations and collections quickly.',
    },
    {
        'icon': 'bi bi-camera-video-fill',
        'title': 'Mentorship & Revision Sessions',
        'description': 'Run online mentor sessions where students revise, practice answers, and track progress scores.',
    },
]
PUBLIC_WORKFLOW = [
    {
        'step': '01',
        'title': 'Create Your Library Workspace',
        'description': 'Enter your library name, open your admin workspace, and start operating with your own identity.',
    },
    {
        'step': '02',
        'title': 'Add Students, Seats, and Plans',
        'description': 'Register students, assign seats by shift, set monthly fees, and track availability instantly.',
    },
    {
        'step': '03',
        'title': 'Run Daily Operations Professionally',
        'description': 'Manage dues, reminders, receipts, reports, and mentor sessions without scattered registers.',
    },
    {
        'step': '04',
        'title': 'Scale the Same Setup for Any Library',
        'description': 'Use the same system for private libraries, study rooms, coaching centers, and campus libraries.',
    },
]
PUBLIC_PRICING = [
    {'name': 'Trial', 'duration': '7 Days', 'price': '49', 'highlight': 'Quick setup for testing the full workflow.'},
    {'name': 'Monthly', 'duration': '1 Month', 'price': '150', 'highlight': 'Best for new libraries starting digital operations.', 'featured': True},
    {'name': 'Quarterly', 'duration': '3 Months', 'price': '450', 'highlight': 'Balanced pricing for growing study spaces.'},
    {'name': 'Annual', 'duration': '1 Year', 'price': '1200', 'highlight': 'Best value for long-term library operations.'},
]
IMAGE_FILE_EXTENSIONS = ('.png', '.jpg', '.jpeg', '.gif', '.webp')


def sorted_students():
    return sorted(Student.objects.all(), key=lambda student: (int(student.seat_number), student.name.lower()))


def file_is_image(file_field):
    return bool(file_field and str(file_field.name).lower().endswith(IMAGE_FILE_EXTENSIONS))


def next_month_slot(year, month):
    if month == 12:
        return year + 1, 1
    return year, month + 1


def build_open_payment_slots(student, months_requested):
    existing_payments = {
        (payment.year, payment.month): payment
        for payment in Payment.objects.filter(student=student)
    }
    year = student.joining_date.year
    month = student.joining_date.month
    slots = []

    for _ in range(240):
        payment = existing_payments.get((year, month))
        if payment is None or not payment.is_paid:
            slots.append((year, month, payment))
            if len(slots) >= months_requested:
                break
        year, month = next_month_slot(year, month)

    return slots


def month_payment_amount(student, months_requested):
    return student.monthly_fee * Decimal(months_requested)


def create_portal_message(student, sender_role, sender_name, body='', attachment=None):
    message_body = (body or '').strip()
    if not message_body and not attachment:
        return None, "Please write a message or attach a payment screenshot."

    portal_message = PortalMessage.objects.create(
        student=student,
        sender_role=sender_role,
        sender_name=sender_name.strip(),
        body=message_body,
        attachment=attachment,
    )
    portal_message.attachment_is_image = file_is_image(portal_message.attachment)
    return portal_message, None


def decorate_portal_messages(message_queryset):
    messages_list = list(message_queryset)
    for portal_message in messages_list:
        portal_message.attachment_is_image = file_is_image(portal_message.attachment)
    return messages_list


def build_portal_conversations():
    conversation_students = list(
        Student.objects.filter(portal_messages__isnull=False)
        .annotate(latest_message_at=Max('portal_messages__created_at'))
        .order_by('-latest_message_at', 'name')
        .distinct()
    )
    latest_message_map = {}

    for portal_message in PortalMessage.objects.select_related('student').order_by('student_id', '-created_at'):
        if portal_message.student_id not in latest_message_map:
            portal_message.attachment_is_image = file_is_image(portal_message.attachment)
            latest_message_map[portal_message.student_id] = portal_message

    for student in conversation_students:
        student.latest_portal_message = latest_message_map.get(student.pk)
        student.portal_message_count = student.portal_messages.count()

    return conversation_students


def razorpay_api_headers():
    credentials = f"{settings.RAZORPAY_KEY_ID}:{settings.RAZORPAY_KEY_SECRET}".encode()
    auth_token = base64.b64encode(credentials).decode()
    return {
        'Authorization': f'Basic {auth_token}',
        'Content-Type': 'application/json',
    }


def create_razorpay_order(student, months_requested):
    if not settings.RAZORPAY_ENABLED:
        raise ValueError("Razorpay is not configured yet.")

    amount = month_payment_amount(student, months_requested)
    amount_paise = int(amount * 100)
    payload = {
        'amount': amount_paise,
        'currency': settings.RAZORPAY_CURRENCY,
        'receipt': f"vl-{student.pk}-{timezone.now().strftime('%Y%m%d%H%M%S')}",
        'notes': {
            'student_name': student.name,
            'seat_number': student.seat_number,
            'months': str(months_requested),
        },
    }
    request_obj = Request(
        'https://api.razorpay.com/v1/orders',
        data=json.dumps(payload).encode(),
        headers=razorpay_api_headers(),
        method='POST',
    )

    try:
        with urlopen(request_obj, timeout=20) as response:
            return json.loads(response.read().decode())
    except HTTPError as exc:
        detail = exc.read().decode(errors='ignore') or exc.reason
        raise ValueError(f"Razorpay order creation failed: {detail}") from exc
    except URLError as exc:
        raise ValueError("Unable to reach Razorpay right now.") from exc


def verify_razorpay_signature(order_id, payment_id, signature):
    expected_signature = hmac.new(
        settings.RAZORPAY_KEY_SECRET.encode(),
        f"{order_id}|{payment_id}".encode(),
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected_signature, signature or '')


@transaction.atomic
def apply_online_payment(student_payment):
    if student_payment.status == 'paid':
        return list(
            Payment.objects.filter(
                student=student_payment.student,
                payment_method='Online',
                paid_on=student_payment.paid_at.date() if student_payment.paid_at else timezone.localdate(),
            ).order_by('year', 'month')
        )

    allocated_payments = []
    for year, month, payment in build_open_payment_slots(student_payment.student, student_payment.months_covered):
        payment = payment or Payment(student=student_payment.student, month=month, year=year)
        payment.is_paid = True
        payment.payment_method = 'Online'
        payment.amount = student_payment.student.monthly_fee
        payment.paid_on = timezone.localdate()
        payment.save()
        allocated_payments.append(payment)

    student_payment.status = 'paid'
    student_payment.paid_at = timezone.now()
    student_payment.notes = ", ".join(payment.month_label for payment in allocated_payments)
    student_payment.save(update_fields=['status', 'paid_at', 'notes', 'updated_at'])
    return allocated_payments


def parse_money(value, fallback):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return fallback


def build_pdf_response(filename, title, lines):
    response = HttpResponse(build_simple_pdf(lines, title=title), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_year_range(today):
    return range(2023, today.year + 3)


def parse_paid_on(value, fallback):
    try:
        return date.fromisoformat(value) if value else fallback
    except ValueError:
        return fallback


def parse_session_time(value):
    parsed = parse_time(value or '')
    return parsed


def parse_session_date(value):
    return parse_date(value or '')


def hydrate_mentorship_session(session):
    session.live_room_url = reverse('mentorship_session_room', args=[session.pk])
    session.has_external_link = bool(session.meeting_link)
    session.show_join_actions = session.join_ready or session.has_external_link
    session.external_meeting_label = (
        'Join Google Meet'
        if session.meeting_link and 'meet.google.com' in session.meeting_link.lower()
        else 'Open Mentor Link'
    )
    return session


def send_library_access_request_email(request, library_name, requester_name):
    send_mail(
        subject=f'New library access request: {library_name}',
        message=(
            "A new library access request was submitted from the public landing page.\n\n"
            f"Library name: {library_name}\n"
            f"Requester name: {requester_name}\n"
            f"Submitted from: {request.get_host()}\n"
            f"Submitted on: {date.today().isoformat()}\n\n"
            "Please create the admin ID/profile for this library and share the access details manually."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[settings.LIBRARY_ACCESS_REQUEST_EMAIL],
        fail_silently=False,
    )


def send_mentorship_join_email(request, mentorship_session):
    if not mentorship_session.student.email:
        raise ValueError("Student email is missing")

    website_room_url = request.build_absolute_uri(
        reverse('mentorship_session_room', args=[mentorship_session.pk])
    )
    schedule_line = "To be confirmed by mentor"
    if mentorship_session.preferred_date and mentorship_session.scheduled_time:
        schedule_line = (
            f"{mentorship_session.preferred_date} at "
            f"{mentorship_session.scheduled_time.strftime('%I:%M %p')}"
        )
    elif mentorship_session.preferred_date:
        schedule_line = str(mentorship_session.preferred_date)

    meet_line = (
        f"Google Meet link: {mentorship_session.meeting_link}\n"
        if mentorship_session.meeting_link else
        "Google Meet link: Not added yet. You can still join from the website room.\n"
    )

    send_mail(
        subject=f"Mentorship session join details: {mentorship_session.topic}",
        message=(
            f"Hello {mentorship_session.student.name},\n\n"
            f"Your mentorship session with {mentorship_session.mentor.name} is ready.\n\n"
            f"Topic: {mentorship_session.topic}\n"
            f"Preparation target: {mentorship_session.preparation_target or mentorship_session.student.preparing_for or 'General revision'}\n"
            f"Scheduled time: {schedule_line}\n\n"
            "Join from the website VC room:\n"
            f"{website_room_url}\n\n"
            f"{meet_line}\n"
            "You can use either the website VC room or the mentor's Google Meet link.\n\n"
            "Valmiki Library Mentorship"
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[mentorship_session.student.email],
        fail_silently=False,
    )


def marketing_home(request):
    if request.method == 'POST':
        library_name = request.POST.get('library', '').strip()
        requester_name = request.POST.get('requester_name', '').strip()
        query = urlencode({'library': library_name, 'requester_name': requester_name})

        if not library_name or not requester_name:
            messages.error(request, 'Please enter both the library name and requester name.')
            return redirect(f"{reverse('marketing_home')}?{query}")

        try:
            send_library_access_request_email(request, library_name, requester_name)
        except Exception:
            messages.error(
                request,
                'The access request could not be sent right now. Please try again after checking email setup.',
            )
            return redirect(f"{reverse('marketing_home')}?{query}")

        messages.success(
            request,
            f"Access request sent for {library_name}. We will review it and create the profile manually.",
        )
        return redirect(f"{reverse('marketing_home')}?{query}")

    return render(request, 'students/public_landing.html', {
        'use_cases': PUBLIC_USE_CASES,
        'features': PUBLIC_FEATURES,
        'workflow_steps': PUBLIC_WORKFLOW,
        'plans': PUBLIC_PRICING,
        'seat_preview_numbers': list(range(21, 37)),
        'library_name_preview': request.GET.get('library', '').strip() or 'Your Library',
        'requester_name_preview': request.GET.get('requester_name', '').strip(),
    })


def student_status_choices():
    return [choice[0] for choice in Student.STATUS_CHOICES]


def student_portal_required(view_func):
    @wraps(view_func)
    def wrapped(request, *args, **kwargs):
        student_id = request.session.get(STUDENT_PORTAL_SESSION_KEY)
        if not student_id:
            return redirect('student_portal_login')

        student = Student.objects.filter(pk=student_id, email_verified=True).first()
        if not student:
            request.session.pop(STUDENT_PORTAL_SESSION_KEY, None)
            return redirect('student_portal_login')

        request.student_portal_student = student
        return view_func(request, *args, **kwargs)

    return wrapped


def mentor_portal_required(view_func):
    @wraps(view_func)
    def wrapped(request, *args, **kwargs):
        mentor_id = request.session.get(MENTOR_PORTAL_SESSION_KEY)
        if not mentor_id:
            return redirect('mentor_portal_login')

        mentor = Mentor.objects.filter(pk=mentor_id, email_verified=True, is_active=True).first()
        if not mentor:
            request.session.pop(MENTOR_PORTAL_SESSION_KEY, None)
            return redirect('mentor_portal_login')

        request.mentor_portal_mentor = mentor
        return view_func(request, *args, **kwargs)

    return wrapped


def build_student_verification_token(student):
    return signing.dumps(
        {'student_id': student.pk, 'email': student.email},
        salt=STUDENT_PORTAL_SIGNING_SALT,
    )


def build_mentor_verification_token(mentor):
    return signing.dumps(
        {'mentor_id': mentor.pk, 'email': mentor.email},
        salt=MENTOR_PORTAL_SIGNING_SALT,
    )


def build_student_verification_url(request, student):
    token = build_student_verification_token(student)
    return request.build_absolute_uri(reverse('student_verify_email', args=[token]))


def build_mentor_verification_url(request, mentor):
    token = build_mentor_verification_token(mentor)
    return request.build_absolute_uri(reverse('mentor_verify_email', args=[token]))


def send_student_verification_email(request, student):
    verification_url = build_student_verification_url(request, student)
    send_mail(
        subject='Verify your Valmiki Library student account',
        message=(
            f"Hello {student.name},\n\n"
            "Please verify your email for the Valmiki Library student portal.\n"
            f"Verification link: {verification_url}\n\n"
            "After verification, you can log in using your name, email, phone number, and date of birth."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[student.email],
        fail_silently=False,
    )
    return verification_url


def send_mentor_verification_email(request, mentor):
    verification_url = build_mentor_verification_url(request, mentor)
    send_mail(
        subject='Verify your Valmiki Library mentor account',
        message=(
            f"Hello {mentor.name},\n\n"
            "Please verify your mentor email for the Valmiki Library mentorship portal.\n"
            f"Verification link: {verification_url}\n\n"
            "After verification, you can log in with your email and phone number to manage mentorship sessions."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[mentor.email],
        fail_silently=False,
    )
    return verification_url


def student_matches_login_details(student, name, email, contact, dob_value):
    return (
        student.name.strip().lower() == name.strip().lower()
        and (student.email or '').strip().lower() == email.strip().lower()
        and (student.contact or '').strip() == contact.strip()
        and student.date_of_birth
        and student.date_of_birth.isoformat() == dob_value
    )


def mentor_matches_login_details(mentor, email, contact):
    return (
        mentor.email.strip().lower() == email.strip().lower()
        and (mentor.contact or '').strip() == contact.strip()
    )


def sorted_mentors():
    return Mentor.objects.filter(is_active=True, email_verified=True).order_by('name')


def mentorship_progress_summary(student):
    sessions = MentorshipSession.objects.filter(student=student).select_related('mentor')
    completed_sessions = [session for session in sessions if session.status == 'Completed']
    marked_sessions = [session for session in completed_sessions if session.marks_awarded is not None]
    average_marks = (
        sum(session.marks_awarded for session in marked_sessions) / len(marked_sessions)
        if marked_sessions else None
    )
    return {
        'total_sessions': sessions.count(),
        'completed_sessions': len(completed_sessions),
        'requested_sessions': sessions.filter(status='Requested').count(),
        'scheduled_sessions': sessions.filter(status='Scheduled').count(),
        'average_marks': average_marks,
    }


def mentor_dashboard_summary(mentor):
    sessions = MentorshipSession.objects.filter(mentor=mentor).select_related('student')
    completed_sessions = sessions.filter(status='Completed')
    marked_sessions = completed_sessions.exclude(marks_awarded__isnull=True)
    average_marks = (
        sum(session.marks_awarded for session in marked_sessions) / marked_sessions.count()
        if marked_sessions.exists() else None
    )
    return {
        'total_sessions': sessions.count(),
        'requested_sessions': sessions.filter(status='Requested').count(),
        'scheduled_sessions': sessions.filter(status='Scheduled').count(),
        'completed_sessions': completed_sessions.count(),
        'active_students': len({session.student_id for session in sessions}),
        'total_revenue': sum((session.price for session in completed_sessions), Decimal('0.00')),
        'average_marks': average_marks,
    }


def save_student_from_request(request, student=None):
    student = student or Student()
    existing_email = (student.email or '').strip().lower()

    name = request.POST['name'].strip()
    contact = request.POST['contact'].strip()
    email = request.POST.get('email', '').strip().lower() or None
    date_of_birth = request.POST.get('date_of_birth') or None
    seat_number = request.POST['seat_number'].strip()
    shift = request.POST['shift']
    joining_date = request.POST['joining_date']
    monthly_fee = request.POST['monthly_fee']
    monthly_due_day = request.POST.get('monthly_due_day') or 5
    mode_of_payment = request.POST['mode_of_payment']
    preparing_for = request.POST.get('preparing_for', '').strip()
    academic_details = request.POST.get('academic_details', '').strip()
    emergency_contact_name = request.POST.get('emergency_contact_name', '').strip()
    emergency_contact_phone = request.POST.get('emergency_contact_phone', '').strip()
    status = request.POST.get('status', 'Active')

    valid_seats = [str(i) for i in range(1, 101)]
    if seat_number not in valid_seats:
        return None, f"❌ Seat {seat_number} invalid"

    if status not in student_status_choices():
        return None, "❌ Invalid student status"

    existing_students = Student.objects.filter(seat_number=seat_number)
    if student.pk:
        existing_students = existing_students.exclude(pk=student.pk)
    shifts = [existing_student.shift for existing_student in existing_students]

    if 'Full Day' in shifts:
        return None, "❌ Seat already occupied for Full Day"
    if shift == 'Full Day' and existing_students.exists():
        return None, "❌ Cannot assign Full Day to occupied seat"
    if len(shifts) >= 2:
        return None, "❌ Seat already has Morning & Evening"
    if shift in shifts:
        return None, f"❌ {shift} already exists on this seat"

    if email:
        email_exists = Student.objects.filter(email__iexact=email)
        if student.pk:
            email_exists = email_exists.exclude(pk=student.pk)
        if email_exists.exists():
            return None, "❌ This email is already used by another student"

    student.name = name
    student.contact = contact
    student.email = email
    student.date_of_birth = date_of_birth
    student.seat_number = seat_number
    student.shift = shift
    student.joining_date = joining_date
    student.monthly_fee = monthly_fee
    student.monthly_due_day = monthly_due_day
    student.mode_of_payment = mode_of_payment
    student.preparing_for = preparing_for
    student.academic_details = academic_details
    student.emergency_contact_name = emergency_contact_name
    student.emergency_contact_phone = emergency_contact_phone
    student.status = status

    if request.FILES.get('profile_photo'):
        student.profile_photo = request.FILES['profile_photo']
    if request.FILES.get('id_copy'):
        student.id_copy = request.FILES['id_copy']
    if request.FILES.get('payment_qr'):
        student.payment_qr = request.FILES['payment_qr']

    if email != existing_email:
        student.email_verified = False

    student.save()
    return student, None


def save_mentor_from_request(request, mentor=None):
    mentor = mentor or Mentor()
    existing_email = (mentor.email or '').strip().lower()

    name = request.POST['name'].strip()
    email = request.POST.get('email', '').strip().lower()
    contact = request.POST.get('contact', '').strip()
    job_role = request.POST.get('job_role', '').strip()
    current_work = request.POST.get('current_work', '').strip()
    experience_years = request.POST.get('experience_years') or 0
    streams_known = request.POST.get('streams_known', '').strip()
    bio = request.POST.get('bio', '').strip()
    primary_session_mode = request.POST.get('primary_session_mode', 'Daily')
    daily_session_price = request.POST.get('daily_session_price') or 0
    monthly_session_price = request.POST.get('monthly_session_price') or 0

    if not email:
        return None, "❌ Mentor email is required"
    if not contact:
        return None, "❌ Mentor phone number is required"
    if not job_role:
        return None, "❌ Mentor job role is required"
    if not streams_known:
        return None, "❌ Please mention the streams or subjects you can mentor"

    email_exists = Mentor.objects.filter(email__iexact=email)
    if mentor.pk:
        email_exists = email_exists.exclude(pk=mentor.pk)
    if email_exists.exists():
        return None, "❌ This mentor email is already registered"

    if primary_session_mode not in [choice[0] for choice in Mentor.SESSION_MODE_CHOICES]:
        return None, "❌ Invalid mentor session mode"

    mentor.name = name
    mentor.email = email
    mentor.contact = contact
    mentor.job_role = job_role
    mentor.current_work = current_work
    mentor.experience_years = experience_years
    mentor.streams_known = streams_known
    mentor.bio = bio
    mentor.primary_session_mode = primary_session_mode
    mentor.daily_session_price = parse_money(daily_session_price, Decimal('0.00'))
    mentor.monthly_session_price = parse_money(monthly_session_price, Decimal('0.00'))

    if request.FILES.get('profile_photo'):
        mentor.profile_photo = request.FILES['profile_photo']

    if email != existing_email:
        mentor.email_verified = False

    mentor.save()
    return mentor, None


def student_form_context(student=None, seat=''):
    return {
        'student_obj': student,
        'seat': student.seat_number if student else seat,
        'status_choices': Student.STATUS_CHOICES,
    }


def mentor_form_context(mentor=None):
    return {
        'mentor_obj': mentor,
        'session_mode_choices': Mentor.SESSION_MODE_CHOICES,
    }


def build_seat_map():
    seat_map = {}
    for student in sorted_students():
        seat_map.setdefault(student.seat_number, {})
        seat_map[student.seat_number][student.shift] = student
    return seat_map


def seat_status_for_shifts(shifts):
    shift_names = set(shifts.keys())
    if 'Full Day' in shift_names:
        return 'full-day'
    if len(shift_names) == 2:
        return 'fully-occupied'
    if len(shift_names) == 1:
        return 'partially-vacant'
    return 'vacant'


def build_seat_overview():
    seat_map = build_seat_map()
    seat_numbers = [str(i) for i in range(1, 101)]
    seat_status_map = {}
    seat_rows = []

    for seat_number in seat_numbers:
        shifts = seat_map.get(seat_number, {})
        status = seat_status_for_shifts(shifts)
        students = sorted(
            shifts.values(),
            key=lambda student: (SHIFT_DISPLAY_ORDER.get(student.shift, 99), student.name.lower()),
        )

        if shifts:
            seat_status_map[seat_number] = status

        status_label = {
            'vacant': 'Vacant',
            'partially-vacant': '1 shift vacant',
            'fully-occupied': 'Occupied',
            'full-day': 'Full Day',
        }[status]

        seat_rows.append({
            'seat_number': seat_number,
            'students': students,
            'status': status,
            'status_label': status_label,
            'is_vacant': status == 'vacant',
            'can_add': status in {'vacant', 'partially-vacant'},
            'is_full': status in {'fully-occupied', 'full-day'},
        })

    return seat_map, seat_numbers, seat_status_map, seat_rows


@login_required
def student_list(request):
    return render(request, 'students/student_list.html', {'students': sorted_students()})


@login_required
def add_student(request):
    error = None

    if request.method == 'POST':
        student, error = save_student_from_request(request)
        if not error:
            messages.success(request, f"{student.name} was added successfully.")
            return redirect('student_list')

    context = student_form_context(seat=request.GET.get('seat', ''))
    context.update({
        'error': error,
        'form_title': '➕ Add New Student',
        'submit_label': 'Save Student →',
        'page_mode': 'add',
    })
    return render(request, 'students/add_student.html', context)


@login_required
def edit_student(request, pk):
    student_obj = get_object_or_404(Student, pk=pk)
    error = None

    if request.method == 'POST':
        updated_student, error = save_student_from_request(request, student=student_obj)
        if not error:
            messages.success(request, f"{updated_student.name}'s profile was updated.")
            return redirect('student_list')

    context = student_form_context(student=student_obj)
    context.update({
        'error': error,
        'form_title': '✏️ Edit Student Profile',
        'submit_label': 'Update Student →',
        'page_mode': 'edit',
    })
    return render(request, 'students/add_student.html', context)


@login_required
def delete_student(request, pk):
    get_object_or_404(Student, pk=pk).delete()
    return redirect('student_list')


@login_required
def seating_chart(request):
    seat_map, seat_numbers, seat_status_map, _ = build_seat_overview()
    return render(request, 'students/seating_chart.html', {
        'seat_map': seat_map,
        'seat_numbers': seat_numbers,
        'seat_status_map': seat_status_map,
    })


@login_required
def student_detail(request):
    _, seat_numbers, seat_status_map, seat_rows = build_seat_overview()
    vacant_seat_count = sum(1 for row in seat_rows if row['is_vacant'])
    partial_seat_count = sum(1 for row in seat_rows if row['status'] == 'partially-vacant')
    occupied_seat_count = len(seat_numbers) - vacant_seat_count
    payment_map = {}
    message_map = {}

    for payment in Payment.objects.filter(is_paid=True).select_related('student').order_by('student_id', '-year', '-month', '-paid_on'):
        receipts = payment_map.setdefault(payment.student_id, [])
        if len(receipts) < 3:
            receipts.append(payment)

    for portal_message in PortalMessage.objects.select_related('student').order_by('student_id', 'created_at'):
        thread = message_map.setdefault(portal_message.student_id, [])
        portal_message.attachment_is_image = file_is_image(portal_message.attachment)
        thread.append(portal_message)

    for row in seat_rows:
        for student in row['students']:
            student.recent_receipts = payment_map.get(student.pk, [])
            student.latest_receipt = student.recent_receipts[0] if student.recent_receipts else None
            student.payment_qr_is_image = file_is_image(student.payment_qr)
            student.portal_message_thread = message_map.get(student.pk, [])[-6:]
            student.portal_message_count = len(message_map.get(student.pk, []))

    return render(request, 'students/student_detail.html', {
        'seat_rows': seat_rows,
        'vacant_seat_count': vacant_seat_count,
        'partial_seat_count': partial_seat_count,
        'occupied_seat_count': occupied_seat_count,
        'seat_status_map': seat_status_map,
        'total_seats': len(seat_numbers),
    })


@login_required
def message_center(request):
    conversations = build_portal_conversations()
    selected_student = None

    selected_student_id = request.GET.get('student')
    if selected_student_id:
        selected_student = Student.objects.filter(pk=selected_student_id).first()
    if selected_student is None and conversations:
        selected_student = conversations[0]

    thread = []
    if selected_student:
        thread = decorate_portal_messages(
            selected_student.portal_messages.order_by('created_at')
        )
        selected_student.payment_qr_is_image = file_is_image(selected_student.payment_qr)

    return render(request, 'students/message_center.html', {
        'conversations': conversations,
        'selected_student': selected_student,
        'thread': thread,
    })


@login_required
def payment_ledger(request):
    today = date.today()
    current_year = int(request.GET.get('year', today.year))
    ledger = [build_ledger_row(student, current_year, today) for student in sorted_students()]

    return render(request, 'students/payment_ledger.html', {
        'ledger': ledger,
        'months': MONTH_NAMES,
        'current_year': current_year,
        'years': build_year_range(today),
    })


@login_required
def payment_entry(request, student_id, month, year):
    if month < 1 or month > 12:
        raise Http404("Invalid month")

    today = date.today()
    student = get_object_or_404(Student, pk=student_id)
    payment = Payment.objects.filter(student=student, month=month, year=year).first()

    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        if action == 'mark_unpaid':
            if payment:
                payment.is_paid = False
                payment.amount = Decimal('0.00')
                payment.paid_on = None
                payment.save(update_fields=['is_paid', 'amount', 'paid_on'])
            return redirect(f'/ledger/?year={year}')

        if payment is None:
            payment = Payment(student=student, month=month, year=year)

        payment.is_paid = True
        payment.payment_method = request.POST.get('payment_method') or student.mode_of_payment or 'Cash'
        payment.amount = parse_money(request.POST.get('amount'), student.monthly_fee)
        payment.paid_on = parse_paid_on(request.POST.get('paid_on'), today)
        payment.save()

        if action == 'save_and_receipt':
            return redirect('payment_receipt_pdf', payment_id=payment.pk)
        return redirect(f'/ledger/?year={year}')

    suggested_method = payment.payment_method if payment and payment.payment_method else student.mode_of_payment
    suggested_amount = payment.amount if payment and payment.amount else student.monthly_fee
    suggested_paid_on = payment.paid_on if payment and payment.paid_on else today

    return render(request, 'students/payment_entry.html', {
        'student': student,
        'month': month,
        'year': year,
        'month_title': month_label(year, month),
        'payment': payment,
        'suggested_method': suggested_method,
        'suggested_amount': suggested_amount,
        'suggested_paid_on': suggested_paid_on,
    })


def payment_receipt_pdf(request, payment_id):
    payment = get_object_or_404(Payment.objects.select_related('student'), pk=payment_id, is_paid=True)
    student_session_id = request.session.get(STUDENT_PORTAL_SESSION_KEY)

    if not request.user.is_authenticated:
        if student_session_id != payment.student_id:
            return redirect('student_portal_login')

    student = payment.student
    lines = [
        f"Receipt Number: {payment.receipt_number}",
        f"Receipt Date: {payment.paid_on or date.today()}",
        "",
        "Received From",
        f"Student Name: {student.name}",
        f"Contact: {student.contact}",
        f"Seat Number: {student.seat_number}",
        f"Shift: {student.shift}",
        "",
        "Payment Details",
        f"Billing Month: {payment.month_label}",
        f"Payment Method: {payment.payment_method}",
        f"Amount Received: Rs. {payment.amount}",
        f"Default Student Mode: {student.mode_of_payment}",
        "",
        "Valmiki Library",
        "This receipt confirms that payment has been received.",
    ]
    filename = f"receipt-{payment.receipt_number}.pdf"
    return build_pdf_response(filename, "Valmiki Library Payment Receipt", lines)


def student_portal_login(request):
    if request.session.get(STUDENT_PORTAL_SESSION_KEY):
        return redirect('student_portal_dashboard')

    error = None
    notice = "Use the same details that are saved in your library profile."
    verification_link = None

    if request.GET.get('verified') == '1':
        notice = "Your email is verified. You can now log in."

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        contact = request.POST.get('contact', '').strip()
        dob_value = request.POST.get('date_of_birth', '')
        action = request.POST.get('action', 'login')

        matching_student = None
        for student in Student.objects.exclude(date_of_birth__isnull=True).exclude(email__isnull=True):
            if student_matches_login_details(student, name, email, contact, dob_value):
                matching_student = student
                break

        if not matching_student:
            error = "No student profile matched these details."
        elif action == 'send_verification':
            verification_link = send_student_verification_email(request, matching_student)
            notice = f"Verification email sent to {matching_student.email}."
        elif not matching_student.email_verified:
            error = "Email is not verified yet. Use the verification button first."
        elif matching_student.status == 'Left':
            error = "This student profile is not active for portal login."
        else:
            request.session[STUDENT_PORTAL_SESSION_KEY] = matching_student.pk
            return redirect('student_portal_dashboard')

    return render(request, 'students/student_portal_login.html', {
        'error': error,
        'notice': notice,
        'verification_link': verification_link if settings.DEBUG else None,
        'show_debug_hint': settings.DEBUG,
    })


def student_verify_email(request, token):
    try:
        payload = signing.loads(token, salt=STUDENT_PORTAL_SIGNING_SALT, max_age=60 * 60 * 24 * 7)
    except signing.BadSignature:
        raise Http404("Invalid or expired verification link")

    student = get_object_or_404(Student, pk=payload.get('student_id'))
    if (student.email or '').strip().lower() != (payload.get('email') or '').strip().lower():
        raise Http404("Verification link does not match student email")

    student.email_verified = True
    student.save(update_fields=['email_verified'])
    return redirect(f"{reverse('student_portal_login')}?verified=1")


def student_portal_logout(request):
    request.session.pop(STUDENT_PORTAL_SESSION_KEY, None)
    return redirect('student_portal_login')


@student_portal_required
def student_portal_dashboard(request):
    student = request.student_portal_student
    today = date.today()
    current_status = get_month_status(student, today.year, today.month, today, paid_lookup(student, today.year))
    overdue = overdue_summary(student, today)
    current_payment = Payment.objects.filter(
        student=student,
        month=today.month,
        year=today.year,
        is_paid=True,
    ).first()
    receipts = Payment.objects.filter(student=student, is_paid=True).order_by('-year', '-month', '-paid_on')[:12]
    ledger_row = build_ledger_row(student, today.year, today)
    mentorship_progress = mentorship_progress_summary(student)
    recent_mentorship_sessions = list(
        MentorshipSession.objects.filter(student=student)
        .select_related('mentor')
        .order_by('-created_at')[:3]
    )
    portal_messages = list(student.portal_messages.order_by('-created_at')[:20])
    portal_messages.reverse()
    decorate_portal_messages(portal_messages)
    for session in recent_mentorship_sessions:
        hydrate_mentorship_session(session)

    return render(request, 'students/student_portal_dashboard.html', {
        'student': student,
        'current_month': month_label(today.year, today.month),
        'current_status': current_status['code'],
        'current_payment': current_payment,
        'overdue': overdue,
        'ledger_row': ledger_row,
        'receipts': receipts,
        'mentorship_progress': mentorship_progress,
        'recent_mentorship_sessions': recent_mentorship_sessions,
        'portal_messages': portal_messages,
        'payment_qr_is_image': file_is_image(student.payment_qr),
        'suggested_online_months': max(1, min(overdue['unpaid_count'] or 1, 12)),
        'razorpay_enabled': settings.RAZORPAY_ENABLED,
        'razorpay_key_id': settings.RAZORPAY_KEY_ID,
    })


@student_portal_required
@require_POST
def student_portal_message(request):
    student = request.student_portal_student
    _, error = create_portal_message(
        student=student,
        sender_role='Student',
        sender_name=student.name,
        body=request.POST.get('body', ''),
        attachment=request.FILES.get('attachment'),
    )

    if error:
        messages.error(request, error)
    else:
        messages.success(request, "Your message was sent to the admin.")
    return redirect('student_portal_dashboard')


@student_portal_required
@require_POST
def student_portal_create_online_payment(request):
    student = request.student_portal_student
    try:
        months_requested = int(request.POST.get('months', '1'))
    except ValueError:
        return JsonResponse({'error': 'Invalid number of months selected.'}, status=400)

    if months_requested < 1 or months_requested > 12:
        return JsonResponse({'error': 'You can pay between 1 and 12 months at a time.'}, status=400)

    try:
        order_data = create_razorpay_order(student, months_requested)
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=503)

    amount = month_payment_amount(student, months_requested)
    purpose = f"Library fee payment for {months_requested} month{'s' if months_requested > 1 else ''}"
    student_payment = StudentOnlinePayment.objects.create(
        student=student,
        amount=amount,
        currency=order_data.get('currency', settings.RAZORPAY_CURRENCY),
        purpose=purpose,
        months_covered=months_requested,
        razorpay_order_id=order_data['id'],
        notes=f"Seat {student.seat_number} · {student.shift}",
    )

    return JsonResponse({
        'key': settings.RAZORPAY_KEY_ID,
        'order_id': student_payment.razorpay_order_id,
        'amount': order_data.get('amount'),
        'currency': student_payment.currency,
        'description': purpose,
        'student_name': student.name,
        'student_email': student.email or '',
        'student_contact': student.contact,
        'seat_number': student.seat_number,
        'library_name': request.GET.get('library') or 'Valmiki Library',
        'months_requested': months_requested,
    })


@student_portal_required
@require_POST
def student_portal_verify_online_payment(request):
    student = request.student_portal_student
    order_id = request.POST.get('razorpay_order_id', '').strip()
    payment_id = request.POST.get('razorpay_payment_id', '').strip()
    signature = request.POST.get('razorpay_signature', '').strip()

    if not order_id or not payment_id or not signature:
        return JsonResponse({'error': 'Payment verification details are incomplete.'}, status=400)

    student_payment = StudentOnlinePayment.objects.filter(
        student=student,
        razorpay_order_id=order_id,
    ).first()
    if not student_payment:
        return JsonResponse({'error': 'Online payment order was not found.'}, status=404)

    if not settings.RAZORPAY_ENABLED:
        return JsonResponse({'error': 'Razorpay is not configured yet.'}, status=503)

    if student_payment.status == 'paid':
        return JsonResponse({
            'success': True,
            'message': 'This online payment was already recorded.',
            'months': student_payment.notes.split(', ') if student_payment.notes else [],
            'receipt_url': '',
        })

    if not verify_razorpay_signature(order_id, payment_id, signature):
        student_payment.status = 'failed'
        student_payment.razorpay_payment_id = payment_id
        student_payment.razorpay_signature = signature
        student_payment.save(update_fields=['status', 'razorpay_payment_id', 'razorpay_signature', 'updated_at'])
        return JsonResponse({'error': 'Payment verification failed. Please contact the admin.'}, status=400)

    student_payment.razorpay_payment_id = payment_id
    student_payment.razorpay_signature = signature
    student_payment.save(update_fields=['razorpay_payment_id', 'razorpay_signature', 'updated_at'])
    allocated_payments = apply_online_payment(student_payment)
    latest_receipt = allocated_payments[-1] if allocated_payments else None

    return JsonResponse({
        'success': True,
        'message': f"Payment recorded successfully for {student_payment.months_covered} month{'s' if student_payment.months_covered > 1 else ''}.",
        'months': [payment.month_label for payment in allocated_payments],
        'receipt_url': reverse('payment_receipt_pdf', args=[latest_receipt.pk]) if latest_receipt else '',
    })


@login_required
@require_POST
def admin_student_message(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    next_url = request.POST.get('next', '').strip()
    _, error = create_portal_message(
        student=student,
        sender_role='Admin',
        sender_name=request.user.get_username() or 'Library Admin',
        body=request.POST.get('body', ''),
        attachment=request.FILES.get('attachment'),
    )

    if error:
        messages.error(request, error)
    else:
        messages.success(request, f"Reply sent to {student.name}.")
    if next_url:
        return redirect(next_url)
    return redirect(f"{reverse('student_detail')}#student-{student.pk}")


def mentor_portal_register(request):
    error = None
    notice = "Create your mentor profile. Verification email will be sent to activate your login."
    verification_link = None

    if request.method == 'POST':
        mentor, error = save_mentor_from_request(request)
        if not error:
            verification_link = send_mentor_verification_email(request, mentor)
            notice = f"Mentor profile created for {mentor.name}. Please verify {mentor.email} before login."

    context = mentor_form_context()
    context.update({
        'error': error,
        'notice': notice,
        'verification_link': verification_link if settings.DEBUG else None,
        'show_debug_hint': settings.DEBUG,
    })
    return render(request, 'students/mentor_portal_register.html', context)


def mentor_portal_login(request):
    if request.session.get(MENTOR_PORTAL_SESSION_KEY):
        return redirect('mentor_portal_dashboard')

    error = None
    notice = "Log in with your verified mentor email and phone number."
    verification_link = None

    if request.GET.get('verified') == '1':
        notice = "Your mentor email is verified. You can now log in."

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        contact = request.POST.get('contact', '').strip()
        action = request.POST.get('action', 'login')

        mentor = Mentor.objects.filter(email__iexact=email).first()
        if not mentor or not mentor_matches_login_details(mentor, email, contact):
            error = "No mentor profile matched these details."
        elif action == 'send_verification':
            verification_link = send_mentor_verification_email(request, mentor)
            notice = f"Verification email sent to {mentor.email}."
        elif not mentor.email_verified:
            error = "Email is not verified yet. Use the verification button first."
        elif not mentor.is_active:
            error = "This mentor profile is inactive."
        else:
            request.session[MENTOR_PORTAL_SESSION_KEY] = mentor.pk
            return redirect('mentor_portal_dashboard')

    return render(request, 'students/mentor_portal_login.html', {
        'error': error,
        'notice': notice,
        'verification_link': verification_link if settings.DEBUG else None,
        'show_debug_hint': settings.DEBUG,
    })


def mentor_verify_email(request, token):
    try:
        payload = signing.loads(token, salt=MENTOR_PORTAL_SIGNING_SALT, max_age=60 * 60 * 24 * 7)
    except signing.BadSignature:
        raise Http404("Invalid or expired verification link")

    mentor = get_object_or_404(Mentor, pk=payload.get('mentor_id'))
    if mentor.email.strip().lower() != (payload.get('email') or '').strip().lower():
        raise Http404("Verification link does not match mentor email")

    mentor.email_verified = True
    mentor.save(update_fields=['email_verified'])
    return redirect(f"{reverse('mentor_portal_login')}?verified=1")


def mentor_portal_logout(request):
    request.session.pop(MENTOR_PORTAL_SESSION_KEY, None)
    return redirect('mentor_portal_login')


@mentor_portal_required
def mentor_portal_dashboard(request):
    mentor = request.mentor_portal_mentor
    summary = mentor_dashboard_summary(mentor)
    sessions = list(
        MentorshipSession.objects.filter(mentor=mentor)
        .select_related('student')
        .order_by('-created_at')
    )
    for session in sessions:
        hydrate_mentorship_session(session)

    return render(request, 'students/mentor_portal_dashboard.html', {
        'mentor': mentor,
        'summary': summary,
        'sessions': sessions,
    })


@mentor_portal_required
def mentor_session_detail(request, session_id):
    mentor = request.mentor_portal_mentor
    mentorship_session = get_object_or_404(
        MentorshipSession.objects.select_related('student', 'mentor'),
        pk=session_id,
        mentor=mentor,
    )
    error = None

    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        status = request.POST.get('status', mentorship_session.status)
        if status not in [choice[0] for choice in MentorshipSession.STATUS_CHOICES]:
            error = "❌ Invalid session status"
        else:
            updated_date = parse_session_date(request.POST.get('preferred_date'))
            updated_time = parse_session_time(request.POST.get('scheduled_time'))
            mentorship_session.status = status
            mentorship_session.preferred_date = updated_date
            mentorship_session.scheduled_time = updated_time
            mentorship_session.meeting_link = request.POST.get('meeting_link', '').strip()
            mentorship_session.mentor_questions = request.POST.get('mentor_questions', '').strip()
            mentorship_session.mentor_feedback = request.POST.get('mentor_feedback', '').strip()
            mentorship_session.price = parse_money(request.POST.get('price'), mentorship_session.price)
            marks_value = request.POST.get('marks_awarded', '').strip()
            mentorship_session.marks_awarded = parse_money(marks_value, mentorship_session.marks_awarded or Decimal('0.00')) if marks_value else None
            if mentorship_session.status == 'Requested' and (
                mentorship_session.meeting_link or mentorship_session.preferred_date or mentorship_session.scheduled_time
            ):
                mentorship_session.status = 'Scheduled'
            mentorship_session.save()

            if action == 'save_and_email':
                try:
                    send_mentorship_join_email(request, mentorship_session)
                except ValueError:
                    messages.warning(
                        request,
                        "Session updated, but the student does not have an email saved yet.",
                    )
                except Exception:
                    messages.error(
                        request,
                        "Session updated, but the join email could not be sent. Check email setup and try again.",
                    )
                else:
                    messages.success(
                        request,
                        f"Mentorship session updated and join details emailed to {mentorship_session.student.email}.",
                    )
            else:
                messages.success(request, "Mentorship session updated.")
            return redirect('mentor_session_detail', session_id=mentorship_session.pk)

    hydrate_mentorship_session(mentorship_session)
    return render(request, 'students/mentor_session_detail.html', {
        'mentor': mentor,
        'mentorship_session': mentorship_session,
        'status_choices': MentorshipSession.STATUS_CHOICES,
        'error': error,
    })


def mentorship_session_room(request, session_id):
    mentorship_session = get_object_or_404(
        MentorshipSession.objects.select_related('student', 'mentor'),
        pk=session_id,
    )
    participant_name = None
    participant_role = None

    if request.user.is_authenticated:
        participant_name = request.user.get_username() or 'Library Admin'
        participant_role = 'Admin Workspace'
    elif request.session.get(STUDENT_PORTAL_SESSION_KEY) == mentorship_session.student_id:
        participant_name = mentorship_session.student.name
        participant_role = 'Student Portal'
    elif request.session.get(MENTOR_PORTAL_SESSION_KEY) == mentorship_session.mentor_id:
        participant_name = mentorship_session.mentor.name
        participant_role = 'Mentor Portal'
    elif request.session.get(MENTOR_PORTAL_SESSION_KEY):
        return redirect('mentor_portal_login')
    else:
        return redirect('student_portal_login')

    hydrate_mentorship_session(mentorship_session)
    return render(request, 'students/mentorship_session_room.html', {
        'mentorship_session': mentorship_session,
        'participant_name': participant_name,
        'participant_role': participant_role,
        'jitsi_room_name': mentorship_session.live_room_name,
        'jitsi_room_url': f"https://meet.jit.si/{mentorship_session.live_room_name}",
    })


@student_portal_required
def student_portal_mentorship(request):
    student = request.student_portal_student
    error = None

    if request.method == 'POST':
        mentor_id = request.POST.get('mentor_id')
        mentor = Mentor.objects.filter(pk=mentor_id, is_active=True, email_verified=True).first()
        plan_type = request.POST.get('plan_type', 'Daily')
        topic = request.POST.get('topic', '').strip()
        study_notes = request.POST.get('student_study_notes', '').strip()
        preparation_target = request.POST.get('preparation_target', '').strip() or student.preparing_for
        preferred_date = request.POST.get('preferred_date') or None

        if not mentor:
            error = "❌ Please choose a valid mentor"
        elif plan_type not in [choice[0] for choice in MentorshipSession.PLAN_CHOICES]:
            error = "❌ Invalid mentorship plan selected"
        elif not topic:
            error = "❌ Please enter the topic or subject you want to practice"
        else:
            price = mentor.daily_session_price if plan_type == 'Daily' else mentor.monthly_session_price
            MentorshipSession.objects.create(
                student=student,
                mentor=mentor,
                plan_type=plan_type,
                topic=topic,
                student_study_notes=study_notes,
                preparation_target=preparation_target,
                preferred_date=preferred_date,
                price=price,
            )
            messages.success(request, f"Mentorship request sent to {mentor.name}.")
            return redirect('student_portal_mentorship')

    sessions = list(
        MentorshipSession.objects.filter(student=student).select_related('mentor').order_by('-created_at')
    )
    for session in sessions:
        hydrate_mentorship_session(session)
    progress = mentorship_progress_summary(student)

    return render(request, 'students/student_portal_mentorship.html', {
        'student': student,
        'mentors': sorted_mentors(),
        'sessions': sessions,
        'progress': progress,
        'plan_choices': MentorshipSession.PLAN_CHOICES,
        'error': error,
        'today': date.today(),
    })


@login_required
def mark_payment(request, student_id, month, year):
    student = get_object_or_404(Student, pk=student_id)
    payment, _ = Payment.objects.get_or_create(student=student, month=month, year=year)
    payment.is_paid = not payment.is_paid
    payment.paid_on = date.today() if payment.is_paid else None
    payment.amount = student.monthly_fee if payment.is_paid else Decimal('0.00')
    payment.payment_method = student.mode_of_payment or 'Cash'
    payment.save()
    return redirect(f'/ledger/?year={year}')


@login_required
def defaulters(request):
    today = date.today()
    defaulter_list = []

    for student in sorted_students():
        current_status = get_month_status(student, today.year, today.month, today, paid_lookup(student, today.year))
        if current_status['code'] != 'unpaid':
            continue

        overdue = overdue_summary(student, today)
        defaulter_list.append({
            'student': student,
            'unpaid_months': overdue['unpaid_count'],
            'total_due': overdue['due_amount'],
        })

    defaulter_list.sort(key=lambda item: item['total_due'], reverse=True)
    return render(request, 'students/defaulters.html', {
        'defaulter_list': defaulter_list,
        'current_month': month_label(today.year, today.month),
        'total_due_all': sum(item['total_due'] for item in defaulter_list),
    })


@login_required
def monthly_reminders(request):
    today = date.today()
    reminder_rows = monthly_reminder_rows(sorted_students(), today)
    return render(request, 'students/monthly_reminders.html', {
        'reminder_rows': reminder_rows,
        'current_month': month_label(today.year, today.month),
        'total_due': sum(row['due_amount'] for row in reminder_rows),
    })


@login_required
def dashboard(request):
    today = date.today()
    students = list(Student.objects.all())
    total_students = len(students)
    occupied_seat_count = len({student.seat_number for student in students})
    vacant_seats = 100 - occupied_seat_count
    occupancy_percent = occupied_seat_count
    morning_count = Student.objects.filter(shift='Morning').count()
    evening_count = Student.objects.filter(shift='Evening').count()
    fullday_count = Student.objects.filter(shift='Full Day').count()

    paid_this_month = Payment.objects.filter(
        month=today.month,
        year=today.year,
        is_paid=True,
    ).count()
    defaulters_count = sum(
        1
        for student in students
        if get_month_status(student, today.year, today.month, today, paid_lookup(student, today.year))['code'] == 'unpaid'
    )

    monthly_earning = Payment.objects.filter(
        month=today.month,
        year=today.year,
        is_paid=True,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    expected_monthly = sum(student.monthly_fee for student in students)
    pending_amount = expected_monthly - monthly_earning

    cash_students = Student.objects.filter(mode_of_payment='Cash').count()
    upi_students = Student.objects.filter(mode_of_payment='UPI').count()
    online_students = Student.objects.filter(mode_of_payment='Online').count()
    recent_students = Student.objects.order_by('-joining_date')[:5]

    return render(request, 'students/dashboard.html', {
        'total_students': total_students,
        'vacant_seats': vacant_seats,
        'occupied_seats': occupied_seat_count,
        'occupancy_percent': occupancy_percent,
        'morning_count': morning_count,
        'evening_count': evening_count,
        'fullday_count': fullday_count,
        'paid_this_month': paid_this_month,
        'defaulters_count': defaulters_count,
        'monthly_earning': monthly_earning,
        'expected_monthly': expected_monthly,
        'pending_amount': pending_amount,
        'cash_students': cash_students,
        'upi_students': upi_students,
        'online_students': online_students,
        'recent_students': recent_students,
        'current_month': month_label(today.year, today.month),
    })


def export_workbook_response(workbook, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def student_directory_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    headers = [
        'Name', 'Contact', 'Email', 'Date of Birth', 'Seat', 'Shift', 'Status',
        'Joining Date', 'Monthly Fee', 'Due Day', 'Payment Mode',
        'Preparing For', 'Emergency Contact', 'Email Verified',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for student in sorted_students():
        sheet.append([
            student.name,
            student.contact,
            student.email or '',
            str(student.date_of_birth or ''),
            student.seat_number,
            student.shift,
            student.status,
            str(student.joining_date),
            float(student.monthly_fee),
            student.monthly_due_day,
            student.mode_of_payment,
            student.preparing_for,
            f"{student.emergency_contact_name} {student.emergency_contact_phone}".strip(),
            'Yes' if student.email_verified else 'No',
        ])

    return workbook


@login_required
def export_students_excel(request):
    return export_workbook_response(student_directory_workbook(), 'students.xlsx')


@login_required
def export_students_pdf(request):
    lines = [
        "Name | Phone | Email | Seat | Shift | Status | Fee | Due Day | Verified",
        "-" * 104,
    ]
    for student in sorted_students():
        lines.append(
            f"{student.name[:14]:<14} | {student.contact:<10} | {(student.email or '-')[:18]:<18} | "
            f"{student.seat_number:<4} | {student.shift[:8]:<8} | {student.status[:7]:<7} | "
            f"Rs. {student.monthly_fee:<6} | {student.monthly_due_day:<7} | "
            f"{'Yes' if student.email_verified else 'No'}"
        )
    return build_pdf_response('students.pdf', 'Valmiki Library Student Directory', lines)


@login_required
def reports_dashboard(request):
    today = date.today()
    selected_year = int(request.GET.get('year', today.year))
    selected_month = int(request.GET.get('month', today.month))
    if selected_month < 1 or selected_month > 12:
        selected_month = today.month
    students = sorted_students()

    monthly_rows, monthly_total_paid, monthly_total_due = monthly_report_rows(
        students, selected_year, selected_month, today
    )
    yearly_rows, yearly_total_paid, yearly_total_due = yearly_report_rows(
        students, selected_year, today
    )

    return render(request, 'students/reports.html', {
        'selected_year': selected_year,
        'selected_month': selected_month,
        'month_title': month_label(selected_year, selected_month),
        'months': list(enumerate(MONTH_NAMES, start=1)),
        'years': build_year_range(today),
        'monthly_preview': monthly_rows[:8],
        'yearly_preview': yearly_rows[:8],
        'monthly_total_paid': monthly_total_paid,
        'monthly_total_due': monthly_total_due,
        'yearly_total_paid': yearly_total_paid,
        'yearly_total_due': yearly_total_due,
    })


def monthly_report_workbook(year, month, today):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Monthly Report"
    sheet.append([f"Valmiki Library Monthly Report - {month_label(year, month)}"])
    sheet.append([])
    headers = ['Name', 'Contact', 'Seat', 'Shift', 'Status', 'Amount Paid', 'Due Amount', 'Payment Method', 'Paid On']
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)

    rows, total_paid, total_due = monthly_report_rows(sorted_students(), year, month, today)
    for row in rows:
        sheet.append([
            row['student'].name,
            row['student'].contact,
            row['student'].seat_number,
            row['student'].shift,
            row['status'],
            float(row['amount_paid']),
            float(row['due_amount']),
            row['payment_method'],
            str(row['paid_on']) if row['paid_on'] else '-',
        ])

    sheet.append([])
    sheet.append(['Total Paid', float(total_paid)])
    sheet.append(['Total Due', float(total_due)])
    return workbook


def yearly_report_workbook(year, today):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Yearly Report"
    sheet.append([f"Valmiki Library Yearly Report - {year}"])
    sheet.append([])
    headers = ['Name', 'Contact', 'Seat', 'Shift', 'Paid Months', 'Advance Months', 'Overdue Months', 'Paid Amount', 'Due Amount']
    sheet.append(headers)
    for cell in sheet[3]:
        cell.font = Font(bold=True)

    rows, total_paid, total_due = yearly_report_rows(sorted_students(), year, today)
    for row in rows:
        sheet.append([
            row['student'].name,
            row['student'].contact,
            row['student'].seat_number,
            row['student'].shift,
            row['paid_months'],
            row['advance_months'],
            row['overdue_months'],
            float(row['paid_amount']),
            float(row['due_amount']),
        ])

    sheet.append([])
    sheet.append(['Total Paid', float(total_paid)])
    sheet.append(['Total Due', float(total_due)])
    return workbook


@login_required
def export_report(request):
    today = date.today()
    scope = request.GET.get('scope', 'monthly')
    export_format = request.GET.get('format', 'xlsx')
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    if month < 1 or month > 12:
        month = today.month

    if scope == 'monthly':
        rows, total_paid, total_due = monthly_report_rows(sorted_students(), year, month, today)
        if export_format == 'xlsx':
            return export_workbook_response(
                monthly_report_workbook(year, month, today),
                f"monthly-report-{year}-{month:02d}.xlsx",
            )

        lines = [
            f"Report Month: {month_label(year, month)}",
            f"Total Paid: Rs. {total_paid}",
            f"Total Due: Rs. {total_due}",
            "",
            "Name | Seat | Shift | Status | Paid | Due | Method | Paid On",
            "-" * 92,
        ]
        for row in rows:
            lines.append(
                f"{row['student'].name[:16]:<16} | {row['student'].seat_number:<4} | "
                f"{row['student'].shift[:8]:<8} | {row['status'][:10]:<10} | "
                f"Rs. {row['amount_paid']:<7} | Rs. {row['due_amount']:<7} | "
                f"{row['payment_method'][:6]:<6} | {row['paid_on'] or '-'}"
            )
        return build_pdf_response(
            f"monthly-report-{year}-{month:02d}.pdf",
            "Valmiki Library Monthly Report",
            lines,
        )

    rows, total_paid, total_due = yearly_report_rows(sorted_students(), year, today)
    if export_format == 'xlsx':
        return export_workbook_response(
            yearly_report_workbook(year, today),
            f"yearly-report-{year}.xlsx",
        )

    lines = [
        f"Report Year: {year}",
        f"Total Paid: Rs. {total_paid}",
        f"Total Due: Rs. {total_due}",
        "",
        "Name | Seat | Shift | Paid | Ahead | Overdue | Paid Amount | Due Amount",
        "-" * 96,
    ]
    for row in rows:
        lines.append(
            f"{row['student'].name[:16]:<16} | {row['student'].seat_number:<4} | "
            f"{row['student'].shift[:8]:<8} | {row['paid_months']:<4} | "
            f"{row['advance_months']:<5} | {row['overdue_months']:<7} | "
            f"Rs. {row['paid_amount']:<7} | Rs. {row['due_amount']}"
        )
    return build_pdf_response(
        f"yearly-report-{year}.pdf",
        "Valmiki Library Yearly Report",
        lines,
    )
